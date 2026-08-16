from flask import (
    abort,
    Blueprint,
    render_template,
    send_file,
    Response
)

from flask_login import (
    login_required,
    current_user
)

from app.models.user import User
from app.models.prediction_history import PredictionHistory
from app.ml.recommend import get_recommendation
from app import db

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


home_bp = Blueprint("home", __name__)


# Landing Page
@home_bp.route("/")
def home():
    return render_template("landing/index.html")


# Medicine Page
@home_bp.route("/medicine")
@login_required
def medicine():

    latest = (
        PredictionHistory.query
        .filter_by(user_id=current_user.id)
        .order_by(PredictionHistory.created_at.desc())
        .first()
    )

    recommendation = None

    if latest:
        recommendation = get_recommendation(latest.disease)

    return render_template(
        "dashboard/medicine.html",
        recommendation=recommendation,
        latest=latest
    )


# Dashboard
@home_bp.route("/dashboard")
@login_required
def dashboard():

    predictions = (
        PredictionHistory.query
        .filter_by(user_id=current_user.id)
        .all()
    )

    total_predictions = len(predictions)

    latest_prediction = (
        predictions[-1]
        if predictions
        else None
    )

    disease_count = {}

    for p in predictions:
        if p.disease:
            disease_count[p.disease] = (
                disease_count.get(p.disease, 0) + 1
            )

    chart_labels = list(disease_count.keys())
    chart_values = list(disease_count.values())

    return render_template(
        "dashboard/dashboard.html",
        total_predictions=total_predictions,
        latest_prediction=latest_prediction,
        chart_labels=chart_labels,
        chart_values=chart_values
    )


# Analytics
@home_bp.route("/analytics")
@login_required
def analytics():

    predictions = (
        PredictionHistory.query
        .filter_by(user_id=current_user.id)
        .all()
    )

    total_predictions = len(predictions)

    disease_count = {}

    for p in predictions:
        if p.disease:
            disease_count[p.disease] = (
                disease_count.get(p.disease, 0) + 1
            )

    most_predicted = None

    if disease_count:
        most_predicted = max(
            disease_count,
            key=disease_count.get
        )

    avg_confidence = 0

    if predictions:
        avg_confidence = round(
            sum(p.confidence for p in predictions)
            / total_predictions,
            2
        )

    return render_template(
        "dashboard/analytics.html",
        total_predictions=total_predictions,
        most_predicted=most_predicted,
        avg_confidence=avg_confidence,
        predictions=predictions
    )

# Doctor Page


@home_bp.route("/doctor")
def doctor():
    return render_template("dashboard/doctor.html")


# Admin Page


@home_bp.route("/admin")
@login_required
def admin():

    if not current_user.is_admin:
        abort(403)

    total_users = User.query.count()
    total_predictions = PredictionHistory.query.count()

    total_diseases = (
        db.session.query(
            PredictionHistory.disease
        ).distinct().count()
    )

    recent_predictions = (
        PredictionHistory.query
        .order_by(PredictionHistory.created_at.desc())
        .limit(10)
        .all()
    )

    return render_template(
        "dashboard/admin.html",
        total_users=total_users,
        total_predictions=total_predictions,
        total_diseases=total_diseases,
        recent_predictions=recent_predictions
    )

# ==========================
# Export PDF
# ==========================


@home_bp.route("/export/pdf")
@login_required
def export_pdf():

    predictions = (
        PredictionHistory.query
        .filter_by(user_id=current_user.id)
        .order_by(PredictionHistory.created_at.desc())
        .all()
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=1.15 * inch,
        bottomMargin=0.85 * inch,
    )

    def _draw_letterhead_and_watermark(canvas, doc):
        """Runs on every page. Draws a text-based letterhead (brand
        name + tagline + rule + page number) and a subtle diagonal
        watermark behind the content.

        Note: this is text-based, not the actual logo artwork —
        ReportLab can't render SVG directly and no SVG-to-PDF
        library (e.g. svglib) is in requirements.txt, so embedding
        the real Logo.svg mark isn't possible without adding a new
        dependency. Flagging this rather than silently working
        around it.
        """

        page_width, page_height = doc.pagesize

        canvas.saveState()

        # --- Letterhead ---
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(colors.HexColor("#0B4F5C"))
        canvas.drawString(
            0.75 * inch, page_height - 0.65 * inch, "Wonderful Crown"
        )

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#5B6672"))
        canvas.drawString(
            0.75 * inch,
            page_height - 0.82 * inch,
            "AI-Assisted Healthcare Recommendation System",
        )

        canvas.setStrokeColor(colors.HexColor("#0B4F5C"))
        canvas.setLineWidth(1)
        canvas.line(
            0.75 * inch,
            page_height - 0.95 * inch,
            page_width - 0.75 * inch,
            page_height - 0.95 * inch,
        )

        # --- Footer: page number ---
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#5B6672"))
        canvas.drawRightString(
            page_width - 0.75 * inch, 0.5 * inch, f"Page {doc.page}"
        )
        canvas.drawString(
            0.75 * inch,
            0.5 * inch,
            "Generated automatically by Wonderful Crown.",
        )

        # --- Watermark: diagonal, low-opacity, behind content ---
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 60)
        canvas.setFillColor(colors.HexColor("#12B3A8"))
        canvas.setFillAlpha(0.06)
        canvas.translate(page_width / 2, page_height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "WONDERFUL CROWN")
        canvas.restoreState()

        canvas.restoreState()

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(
        "<b><font size='20'>Wonderful Crown Prediction Report</font></b>",
        styles["Title"]
    )

    elements.append(title)
    elements.append(Spacer(1, 20))

    intro = Paragraph(
        f"""
        This report contains the prediction history generated by
        <b>{current_user.name}</b> using the Wonderful Crown Healthcare
        Recommendation System.
        """,
        styles["BodyText"]
    )

    elements.append(intro)
    elements.append(Spacer(1, 20))

    table_data = [
        [
            "Disease",
            "Confidence",
            "Prediction Date"
        ]
    ]

    for p in predictions:

        table_data.append([
            p.disease.replace("_", " ").title(),
            f"{p.confidence:.2f} %",
            p.created_at.strftime("%d-%m-%Y")
        ])

    table = Table(table_data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B4F5C")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

            ("GRID", (0, 0), (-1, -1), 1, colors.grey),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ])
    )

    elements.append(table)

    elements.append(Spacer(1, 20))

    document.build(
        elements,
        onFirstPage=_draw_letterhead_and_watermark,
        onLaterPages=_draw_letterhead_and_watermark,
    )

    buffer.seek(0)

    return send_file(
        buffer,
        download_name="WonderfulCrown_Report.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )

# ==========================
# Export Excel
# ==========================


@home_bp.route("/export/excel")
@login_required
def export_excel():

    predictions = (
        PredictionHistory.query
        .filter_by(user_id=current_user.id)
        .order_by(PredictionHistory.created_at.desc())
        .all()
    )

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Prediction Report"

    headers = [
        "Disease",
        "Confidence (%)",
        "Prediction Date"
    ]

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    for col, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=col)

        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    row = 2

    for p in predictions:

        sheet.cell(row=row, column=1).value = (
            p.disease.replace("_", " ").title()
        )

        sheet.cell(row=row, column=2).value = (
            p.confidence
        )

        sheet.cell(row=row, column=3).value = (
            p.created_at.strftime("%d-%m-%Y")
        )

        row += 1

    for column in sheet.columns:

        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column
        )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = length + 5

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="WonderfulCrown_Report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@home_bp.route("/make-admin")
def make_admin():

    user = User.query.filter_by(
        email="jiten282003@gmail.com"
    ).first()

    if user:
        user.is_admin = True
        db.session.commit()
        return "Admin Created Successfully"

    return "User not found"
